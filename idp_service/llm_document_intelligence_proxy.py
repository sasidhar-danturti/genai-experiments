"""LLM-backed proxy that mimics Azure Document Intelligence responses."""

from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass
from email import message_from_bytes
from email.message import Message
from typing import Any, Dict, Iterable, List, Optional

try:  # pragma: no cover - optional dependency for PDF parsing
    import fitz  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    fitz = None  # type: ignore

try:  # pragma: no cover - optional dependency for Excel parsing
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    openpyxl = None  # type: ignore

logger = logging.getLogger(__name__)


DEFAULT_CONFIDENCE = 0.93


class _ProxyAnalyzePoller:
    """Simple poller that mirrors the Azure SDK poller contract."""

    def __init__(self, result: Dict[str, Any]):
        self._result = {"analyzeResult": result}

    def result(self) -> Dict[str, Any]:
        return self._result


@dataclass
class ProxyDocument:
    """Parsed document artefacts used to mimic Azure responses."""

    paragraphs: List[Dict[str, Any]]
    tables: List[Dict[str, Any]]
    documents: List[Dict[str, Any]]
    pages: List[Dict[str, Any]]


class LLMAzureDocumentIntelligenceClient:
    """Light-weight client that mimics Azure Document Intelligence outputs."""

    def __init__(self, *, default_doc_type: str = "general") -> None:
        self.default_doc_type = default_doc_type

    # ------------------------------------------------------------------
    # Azure-compatible interface
    # ------------------------------------------------------------------
    def begin_analyze_document(
        self,
        model_id: str,
        document: Any,
        **kwargs: Any,
    ) -> _ProxyAnalyzePoller:
        document_bytes = _coerce_bytes(document)
        if not document_bytes:
            raise ValueError("Document payload must be bytes-like")

        content_type = kwargs.get("content_type")
        pages = kwargs.get("pages")

        parsed = self._parse_document(document_bytes, content_type=content_type, pages=pages)
        payload = {
            "modelId": model_id,
            "paragraphs": parsed.paragraphs,
            "tables": parsed.tables,
            "documents": parsed.documents,
            "pages": parsed.pages,
        }
        return _ProxyAnalyzePoller(payload)

    # ------------------------------------------------------------------
    # Parsing helpers
    # ------------------------------------------------------------------
    def _parse_document(
        self,
        payload: bytes,
        *,
        content_type: Optional[str],
        pages: Optional[Iterable[int]],
    ) -> ProxyDocument:
        mime = (content_type or "").lower()
        if _looks_like_email(mime, payload):
            return self._parse_email(payload)
        if _looks_like_csv(mime, payload):
            return self._parse_csv(payload)
        if _looks_like_excel(mime, payload):
            return self._parse_excel(payload)
        if _looks_like_pdf(mime, payload):
            return self._parse_pdf(payload, pages=pages)

        # Default text handler
        return self._parse_text(payload)

    # ------------------------------------------------------------------
    # Concrete parsers
    # ------------------------------------------------------------------
    def _parse_pdf(
        self,
        payload: bytes,
        *,
        pages: Optional[Iterable[int]],
    ) -> ProxyDocument:
        if fitz is None:
            logger.warning("PyMuPDF is unavailable; falling back to text heuristics for PDF")
            return self._parse_text(payload)

        document = fitz.open(stream=payload, filetype="pdf")
        selected_pages = {int(p) for p in pages} if pages else None

        paragraphs: List[Dict[str, Any]] = []
        tables: List[Dict[str, Any]] = []
        pages_payload: List[Dict[str, Any]] = []

        try:
            for index in range(document.page_count):
                page_number = index + 1
                if selected_pages and page_number not in selected_pages:
                    continue

                page = document.load_page(index)
                pages_payload.append({"pageNumber": page_number, "confidence": DEFAULT_CONFIDENCE})

                text_dict = page.get_text("dict")
                for block_index, block in enumerate(text_dict.get("blocks", [])):
                    if block.get("type") != 0:
                        continue
                    block_text = _normalise_text(
                        "\n".join(
                            span.get("text", "")
                            for line in block.get("lines", [])
                            for span in line.get("spans", [])
                        )
                    )
                    if not block_text:
                        continue
                    bbox = block.get("bbox") or [0, 0, 0, 0]
                    paragraphs.append(
                        {
                            "content": block_text,
                            "id": f"page-{page_number}-block-{block_index}",
                            "confidence": DEFAULT_CONFIDENCE,
                            "boundingRegions": [
                                {
                                    "pageNumber": page_number,
                                    "boundingBox": list(bbox),
                                }
                            ],
                        }
                    )

                # Basic table heuristics using words arranged in rectangular grids
                if "tables" in text_dict:
                    tables.extend(_coerce_tables_from_dict(text_dict, page_number))
        finally:
            document.close()

        if not pages_payload:
            pages_payload = [{"pageNumber": 1, "confidence": DEFAULT_CONFIDENCE}]

        documents = [
            {
                "docType": self.default_doc_type,
                "fields": {
                    "PageCount": {
                        "value": len(pages_payload),
                        "type": "number",
                        "confidence": DEFAULT_CONFIDENCE,
                    }
                },
            }
        ]

        return ProxyDocument(paragraphs=paragraphs, tables=tables, documents=documents, pages=pages_payload)

    def _parse_csv(self, payload: bytes) -> ProxyDocument:
        text = payload.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        rows = [list(row) for row in reader]
        if not rows:
            rows = [[]]

        header = rows[0] if rows else []
        paragraphs = [
            {
                "content": ", ".join(header),
                "id": "csv-header",
                "confidence": DEFAULT_CONFIDENCE,
                "boundingRegions": [{"pageNumber": 1}],
            }
        ]

        cells = []
        for row_index, row in enumerate(rows):
            for column_index, value in enumerate(row):
                cells.append(
                    {
                        "rowIndex": row_index,
                        "columnIndex": column_index,
                        "content": value,
                        "confidence": DEFAULT_CONFIDENCE,
                        "boundingRegions": [{"pageNumber": 1}],
                    }
                )

        tables = [
            {
                "id": "csv-table",
                "confidence": DEFAULT_CONFIDENCE,
                "cells": cells,
                "caption": "CSV data table",
            }
        ]

        documents = [
            {
                "docType": "tabular",
                "fields": {
                    "RowCount": {
                        "value": len(rows),
                        "type": "number",
                        "confidence": DEFAULT_CONFIDENCE,
                    }
                },
            }
        ]

        pages = [{"pageNumber": 1, "confidence": DEFAULT_CONFIDENCE}]
        return ProxyDocument(paragraphs=paragraphs, tables=tables, documents=documents, pages=pages)

    def _parse_excel(self, payload: bytes) -> ProxyDocument:
        if openpyxl is None:
            logger.warning("openpyxl is unavailable; falling back to CSV heuristics for Excel")
            return self._parse_csv(payload)

        workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
        sheet = workbook.active

        paragraphs: List[Dict[str, Any]] = []
        cells = []
        for row in sheet.iter_rows(values_only=True):
            if not paragraphs and row:
                paragraphs.append(
                    {
                        "content": ", ".join(str(value) for value in row if value is not None),
                        "id": "excel-header",
                        "confidence": DEFAULT_CONFIDENCE,
                        "boundingRegions": [{"pageNumber": 1}],
                    }
                )
            break

        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            for column_index, value in enumerate(row):
                cells.append(
                    {
                        "rowIndex": row_index,
                        "columnIndex": column_index,
                        "content": "" if value is None else str(value),
                        "confidence": DEFAULT_CONFIDENCE,
                        "boundingRegions": [{"pageNumber": 1}],
                    }
                )

        tables = [
            {
                "id": f"sheet-{sheet.title}",
                "confidence": DEFAULT_CONFIDENCE,
                "cells": cells,
                "caption": f"Sheet {sheet.title}",
            }
        ]

        documents = [
            {
                "docType": "spreadsheet",
                "fields": {
                    "SheetName": {
                        "value": sheet.title,
                        "type": "string",
                        "confidence": DEFAULT_CONFIDENCE,
                    },
                    "RowCount": {
                        "value": sheet.max_row,
                        "type": "number",
                        "confidence": DEFAULT_CONFIDENCE,
                    },
                },
            }
        ]

        pages = [{"pageNumber": 1, "confidence": DEFAULT_CONFIDENCE}]
        return ProxyDocument(paragraphs=paragraphs, tables=tables, documents=documents, pages=pages)

    def _parse_email(self, payload: bytes) -> ProxyDocument:
        message = message_from_bytes(payload)
        if not isinstance(message, Message):
            raise ValueError("Unable to parse email payload")

        body_text = _extract_email_body(message)
        paragraphs = [
            {
                "content": body_text,
                "id": "email-body",
                "confidence": DEFAULT_CONFIDENCE,
                "boundingRegions": [{"pageNumber": 1}],
            }
        ]

        fields = {
            "Subject": {
                "value": message.get("Subject"),
                "type": "string",
                "confidence": DEFAULT_CONFIDENCE,
            },
            "From": {
                "value": message.get("From"),
                "type": "string",
                "confidence": DEFAULT_CONFIDENCE,
            },
            "To": {
                "value": message.get("To"),
                "type": "string",
                "confidence": DEFAULT_CONFIDENCE,
            },
        }

        attachment_fields = []
        for attachment_index, attachment in enumerate(_iter_attachments(message)):
            attachment_fields.append(
                {
                    "value": attachment.get("filename"),
                    "type": "string",
                    "confidence": DEFAULT_CONFIDENCE,
                    "modelId": attachment.get("content_type"),
                    "boundingRegions": [{"pageNumber": 1}],
                }
            )
        if attachment_fields:
            fields["Attachments"] = {
                "value": json.dumps([field.get("value") for field in attachment_fields]),
                "type": "list",
                "confidence": DEFAULT_CONFIDENCE,
            }

        documents = [
            {
                "docType": "email",
                "fields": fields,
            }
        ]

        pages = [{"pageNumber": 1, "confidence": DEFAULT_CONFIDENCE}]
        return ProxyDocument(paragraphs=paragraphs, tables=[], documents=documents, pages=pages)

    def _parse_text(self, payload: bytes) -> ProxyDocument:
        text = _normalise_text(payload.decode("utf-8", errors="ignore"))
        if not text:
            text = "(empty document)"

        paragraphs = [
            {
                "content": text,
                "id": "text-body",
                "confidence": DEFAULT_CONFIDENCE,
                "boundingRegions": [{"pageNumber": 1}],
            }
        ]

        documents = [
            {
                "docType": self.default_doc_type,
                "fields": {
                    "CharacterCount": {
                        "value": len(text),
                        "type": "number",
                        "confidence": DEFAULT_CONFIDENCE,
                    }
                },
            }
        ]

        pages = [{"pageNumber": 1, "confidence": DEFAULT_CONFIDENCE}]
        return ProxyDocument(paragraphs=paragraphs, tables=[], documents=documents, pages=pages)


# ----------------------------------------------------------------------
# Helper utilities
# ----------------------------------------------------------------------


def _coerce_bytes(value: Any) -> bytes:
    if isinstance(value, bytes):
        return value
    if hasattr(value, "read"):
        return value.read()
    if isinstance(value, str):
        return value.encode("utf-8")
    raise TypeError("Document payload must be bytes-like or a file object")


def _looks_like_pdf(mime: str, payload: bytes) -> bool:
    if mime.startswith("application/pdf"):
        return True
    return payload[:4] == b"%PDF"


def _looks_like_email(mime: str, payload: bytes) -> bool:
    if mime in {"message/rfc822", "application/vnd.ms-outlook"}:
        return True
    return payload.startswith(b"From:") or b"Content-Type: multipart" in payload[:200]


def _looks_like_csv(mime: str, payload: bytes) -> bool:
    if mime.endswith("csv"):
        return True
    sample = payload[:100]
    return b"," in sample and b"\n" in sample


def _looks_like_excel(mime: str, payload: bytes) -> bool:
    if mime in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }:
        return True
    # XLSX files start with PK due to zip
    return payload[:2] == b"PK"


def _normalise_text(value: str) -> str:
    clean = re.sub(r"\s+", " ", value or "").strip()
    return clean


def _coerce_tables_from_dict(text_dict: Dict[str, Any], page_number: int) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    tables = text_dict.get("tables") or []
    for table_index, table in enumerate(tables):
        cells = []
        for cell in table.get("cells", []):
            text = _normalise_text(cell.get("text") or cell.get("content") or "")
            cells.append(
                {
                    "rowIndex": cell.get("row", cell.get("rowIndex", 0)),
                    "columnIndex": cell.get("col", cell.get("columnIndex", 0)),
                    "content": text,
                    "confidence": DEFAULT_CONFIDENCE,
                    "boundingRegions": [
                        {
                            "pageNumber": page_number,
                            "boundingBox": list(cell.get("bbox") or cell.get("rect") or []),
                        }
                    ],
                }
            )
        results.append(
            {
                "id": f"pdf-table-{page_number}-{table_index}",
                "confidence": DEFAULT_CONFIDENCE,
                "cells": cells,
                "caption": table.get("caption"),
            }
        )
    return results


def _extract_email_body(message: Message) -> str:
    if message.is_multipart():
        for part in message.walk():
            content_type = part.get_content_type()
            if content_type == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    return payload.decode(part.get_content_charset() or "utf-8", errors="ignore")
        # Fallback to HTML converted to text
        for part in message.walk():
            if part.get_content_type() == "text/html":
                payload = part.get_payload(decode=True)
                if payload:
                    text = re.sub("<[^>]+>", " ", payload.decode(part.get_content_charset() or "utf-8", errors="ignore"))
                    return _normalise_text(text)
    payload = message.get_payload(decode=True)
    if not payload:
        return ""
    charset = message.get_content_charset() or "utf-8"
    return payload.decode(charset, errors="ignore")


def _iter_attachments(message: Message) -> Iterable[Dict[str, Any]]:
    for part in message.walk():
        if part.get_content_disposition() == "attachment":
            payload = part.get_payload(decode=True) or b""
            yield {
                "filename": part.get_filename(),
                "content_type": part.get_content_type(),
                "payload": payload,
            }

